#!/usr/bin/env python3
"""
agent_d.py — Phase 1: Discover and report Azure cloud resources

Skill: discovery-azure-cloud v1.0.0
Author: skill-advantage

Discovers all resource groups, all resources (running and stopped) in an Azure
subscription, retrieves current-month cost data, and exports everything to an
Excel workbook.

Usage:
    python agent_d.py --confirm

Environment variables required:
    AZURE_SUBSCRIPTION_ID   — Azure subscription ID
    AZURE_TENANT_ID         — Azure AD tenant ID
    AZURE_CLIENT_ID         — Service principal / app registration client ID
    AZURE_CLIENT_SECRET     — Service principal client secret

Optional:
    AZURE_OUTPUT_PATH       — Output Excel file path (default: azure_discovery.xlsx)

Exit codes:
    0 — success
    1 — configuration / credential error
    2 — API / SDK error
    3 — missing --confirm flag
"""

import argparse
import logging
import os
import sys
import time
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

# ---------------------------------------------------------------------------
# Dependency check — give a clear message before crashing on import errors
# ---------------------------------------------------------------------------
_MISSING: List[str] = []
try:
    from azure.identity import ClientSecretCredential
except ImportError:
    _MISSING.append("azure-identity")
try:
    from azure.mgmt.resource import ResourceManagementClient
except ImportError:
    _MISSING.append("azure-mgmt-resource")
try:
    from azure.mgmt.costmanagement import CostManagementClient
    from azure.mgmt.costmanagement.models import (
        QueryDefinition,
        QueryTimePeriod,
        QueryDataset,
        QueryAggregation,
        QueryGrouping,
        TimeframeType,
        ExportType,
    )
except ImportError:
    _MISSING.append("azure-mgmt-costmanagement")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    _MISSING.append("openpyxl")

if _MISSING:
    print(
        f"[ERROR] Missing required packages: {', '.join(_MISSING)}\n"
        f"        Install with: pip install {' '.join(_MISSING)}",
        file=sys.stderr,
    )
    sys.exit(1)

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%dT%H:%M:%S",
)
log = logging.getLogger("azure-discovery")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
MAX_RETRIES = 4
RETRY_BACKOFF_BASE = 2  # seconds — exponential backoff


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _retry(fn, *args, label: str = "API call", **kwargs) -> Any:
    """Call *fn* with exponential back-off on transient errors."""
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            return fn(*args, **kwargs)
        except Exception as exc:  # noqa: BLE001
            wait = RETRY_BACKOFF_BASE ** attempt
            if attempt < MAX_RETRIES:
                log.warning(
                    "%s failed (attempt %d/%d): %s — retrying in %ds",
                    label, attempt, MAX_RETRIES, exc, wait,
                )
                time.sleep(wait)
            else:
                log.error("%s failed after %d attempts: %s", label, MAX_RETRIES, exc)
                raise


def _env(name: str, required: bool = True) -> Optional[str]:
    val = os.environ.get(name, "").strip()
    if required and not val:
        log.error("Required environment variable '%s' is not set.", name)
        sys.exit(1)
    return val or None


# ---------------------------------------------------------------------------
# Azure clients
# ---------------------------------------------------------------------------

def build_credential() -> ClientSecretCredential:
    tenant_id = _env("AZURE_TENANT_ID")
    client_id = _env("AZURE_CLIENT_ID")
    client_secret = _env("AZURE_CLIENT_SECRET")
    log.info("Authenticating to Azure AD tenant %s …", tenant_id)
    return ClientSecretCredential(
        tenant_id=tenant_id,
        client_id=client_id,
        client_secret=client_secret,
    )


# ---------------------------------------------------------------------------
# Discovery functions
# ---------------------------------------------------------------------------

def list_resource_groups(
    client: ResourceManagementClient,
) -> List[Dict[str, Any]]:
    log.info("Listing resource groups …")
    groups: List[Dict[str, Any]] = []
    for rg in _retry(client.resource_groups.list, label="list_resource_groups"):
        groups.append(
            {
                "name": rg.name,
                "location": rg.location,
                "provisioning_state": rg.properties.provisioning_state
                if rg.properties
                else "Unknown",
                "tags": str(rg.tags or {}),
            }
        )
    log.info("Found %d resource group(s).", len(groups))
    return groups


def list_resources(
    client: ResourceManagementClient,
    subscription_id: str,
) -> List[Dict[str, Any]]:
    """List every resource in the subscription regardless of power-state."""
    log.info("Listing all resources in subscription %s …", subscription_id)
    resources: List[Dict[str, Any]] = []

    # expand=provisioningState gives per-resource state; createdTime/changedTime
    # require a separate call on some resource types — we include what the API
    # returns and fall back to 'N/A'.
    raw_resources = list(
        _retry(
            client.resources.list,
            expand="createdTime,changedTime,provisioningState",
            label="list_resources",
        )
    )

    log.info("Raw resource count: %d", len(raw_resources))
    for res in raw_resources:
        # Determine an operational status where available
        prov_state = "Unknown"
        if res.provisioning_state:
            prov_state = res.provisioning_state

        resources.append(
            {
                "id": res.id or "",
                "name": res.name or "",
                "type": res.type or "",
                "resource_group": _extract_rg(res.id or ""),
                "location": res.location or "global",
                "sku": _format_sku(res.sku),
                "kind": res.kind or "",
                "provisioning_state": prov_state,
                "created_time": str(res.created_time) if res.created_time else "N/A",
                "changed_time": str(res.changed_time) if res.changed_time else "N/A",
                "tags": str(res.tags or {}),
            }
        )

    log.info("Total resources discovered: %d", len(resources))
    return resources


def _extract_rg(resource_id: str) -> str:
    """Extract resource-group name from an Azure resource ID."""
    parts = resource_id.lower().split("/")
    try:
        idx = parts.index("resourcegroups")
        return resource_id.split("/")[idx + 1]
    except (ValueError, IndexError):
        return "Unknown"


def _format_sku(sku) -> str:
    if sku is None:
        return ""
    parts = []
    if sku.name:
        parts.append(sku.name)
    if sku.tier:
        parts.append(sku.tier)
    if sku.size:
        parts.append(sku.size)
    return " / ".join(parts)


# ---------------------------------------------------------------------------
# Cost Management
# ---------------------------------------------------------------------------

def get_current_month_costs(
    cost_client: CostManagementClient,
    subscription_id: str,
) -> Tuple[float, str, List[Dict[str, Any]]]:
    """
    Return (total_cost, currency, per_service_rows).

    Uses the Cost Management 'usage' query scoped to the current billing month.
    Falls back gracefully if the subscription has no cost data yet.
    """
    scope = f"/subscriptions/{subscription_id}"
    now = datetime.now(tz=timezone.utc)
    first_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    log.info(
        "Querying cost management for %s → %s …",
        first_of_month.date(),
        now.date(),
    )

    query_def = QueryDefinition(
        type=ExportType.ACTUAL_COST,
        timeframe=TimeframeType.CUSTOM,
        time_period=QueryTimePeriod(
            from_property=first_of_month,
            to=now,
        ),
        dataset=QueryDataset(
            granularity="None",
            aggregation={
                "totalCost": QueryAggregation(name="Cost", function="Sum"),
            },
            grouping=[
                QueryGrouping(type="Dimension", name="ServiceName"),
            ],
        ),
    )

    try:
        result = _retry(
            cost_client.query.usage,
            scope=scope,
            parameters=query_def,
            label="cost_query",
        )
    except Exception as exc:  # noqa: BLE001
        log.warning("Cost query failed: %s — cost data will be unavailable.", exc)
        return 0.0, "N/A", []

    rows: List[Dict[str, Any]] = []
    total = 0.0
    currency = "USD"

    if result.rows:
        # Columns vary; locate cost and service name by column metadata
        col_names = [c.name for c in (result.columns or [])]
        cost_idx = _find_col(col_names, "Cost", "PreTaxCost", "TotalCost")
        svc_idx = _find_col(col_names, "ServiceName")
        cur_idx = _find_col(col_names, "Currency")

        for row in result.rows:
            cost_val = float(row[cost_idx]) if cost_idx is not None else 0.0
            svc_name = str(row[svc_idx]) if svc_idx is not None else "Unknown"
            if cur_idx is not None:
                currency = str(row[cur_idx])
            total += cost_val
            rows.append({"service_name": svc_name, "cost": round(cost_val, 4), "currency": currency})

    log.info("Total current-month cost: %.2f %s across %d services.", total, currency, len(rows))
    return round(total, 2), currency, rows


def _find_col(col_names: List[str], *candidates: str) -> Optional[int]:
    for candidate in candidates:
        for i, name in enumerate(col_names):
            if name.lower() == candidate.lower():
                return i
    return None


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

# Colour palette
HEADER_FILL_BLUE = PatternFill("solid", fgColor="1F4E79")
HEADER_FILL_GREEN = PatternFill("solid", fgColor="375623")
HEADER_FILL_PURPLE = PatternFill("solid", fgColor="4B0082")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
ALT_FILL = PatternFill("solid", fgColor="DEEAF1")
BORDER_SIDE = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(
    left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE
)


def _write_header_row(ws, headers: List[str], fill: PatternFill, row: int = 1):
    for col, heading in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=heading)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _write_data_rows(ws, data: List[List[Any]], start_row: int = 2):
    for r_idx, row_data in enumerate(data):
        fill = ALT_FILL if r_idx % 2 == 0 else PatternFill()
        for c_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=start_row + r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if fill.fill_type:
                cell.fill = fill


def _auto_fit_columns(ws, min_width: int = 12, max_width: int = 60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:  # noqa: BLE001
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def generate_excel(
    subscription_id: str,
    resource_groups: List[Dict[str, Any]],
    resources: List[Dict[str, Any]],
    total_cost: float,
    currency: str,
    cost_rows: List[Dict[str, Any]],
    output_path: str,
):
    log.info("Generating Excel workbook: %s …", output_path)
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1 — Summary
    # ------------------------------------------------------------------ #
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.sheet_view.showGridLines = False

    ws_summary["A1"] = "Azure Subscription Discovery Report"
    ws_summary["A1"].font = TITLE_FONT
    ws_summary.merge_cells("A1:D1")

    summary_data = [
        ["Subscription ID", subscription_id],
        ["Report Generated (UTC)", datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S")],
        ["Total Resource Groups", len(resource_groups)],
        ["Total Resources", len(resources)],
        [
            "Running / Succeeded Resources",
            sum(
                1 for r in resources
                if r["provisioning_state"].lower() in ("succeeded", "running", "ready")
            ),
        ],
        [
            "Non-running Resources",
            sum(
                1 for r in resources
                if r["provisioning_state"].lower() not in ("succeeded", "running", "ready")
            ),
        ],
        [f"Total Current-Month Cost ({currency})", total_cost],
    ]

    for row_idx, (label, value) in enumerate(summary_data, start=3):
        ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2, value=value)

    _auto_fit_columns(ws_summary)

    # ------------------------------------------------------------------ #
    # Sheet 2 — Resource Groups
    # ------------------------------------------------------------------ #
    ws_rg = wb.create_sheet("Resource Groups")
    ws_rg.sheet_view.showGridLines = False

    rg_headers = ["Name", "Location", "Provisioning State", "Tags"]
    _write_header_row(ws_rg, rg_headers, HEADER_FILL_BLUE)

    rg_data = [
        [rg["name"], rg["location"], rg["provisioning_state"], rg["tags"]]
        for rg in resource_groups
    ]
    _write_data_rows(ws_rg, rg_data)
    ws_rg.auto_filter.ref = f"A1:D{len(rg_data) + 1}"
    _auto_fit_columns(ws_rg)

    # ------------------------------------------------------------------ #
    # Sheet 3 — All Resources
    # ------------------------------------------------------------------ #
    ws_res = wb.create_sheet("All Resources")
    ws_res.sheet_view.showGridLines = False
    ws_res.row_dimensions[1].height = 30

    res_headers = [
        "Name", "Resource Group", "Type", "Location", "SKU",
        "Kind", "Provisioning State", "Created Time", "Changed Time", "Tags",
    ]
    _write_header_row(ws_res, res_headers, HEADER_FILL_GREEN)

    res_data = [
        [
            r["name"], r["resource_group"], r["type"], r["location"],
            r["sku"], r["kind"], r["provisioning_state"],
            r["created_time"], r["changed_time"], r["tags"],
        ]
        for r in resources
    ]
    _write_data_rows(ws_res, res_data)
    ws_res.auto_filter.ref = f"A1:{get_column_letter(len(res_headers))}{len(res_data) + 1}"
    _auto_fit_columns(ws_res)

    # ------------------------------------------------------------------ #
    # Sheet 4 — Cost by Service
    # ------------------------------------------------------------------ #
    ws_cost = wb.create_sheet("Cost by Service")
    ws_cost.sheet_view.showGridLines = False

    cost_headers = ["Service Name", f"Cost ({currency})", "Currency"]
    _write_header_row(ws_cost, cost_headers, HEADER_FILL_PURPLE)

    cost_data_rows = [
        [c["service_name"], c["cost"], c["currency"]] for c in cost_rows
    ]
    _write_data_rows(ws_cost, cost_data_rows)

    # Total row
    total_row = len(cost_data_rows) + 2
    total_cell = ws_cost.cell(row=total_row, column=1, value="TOTAL")
    total_cell.font = Font(bold=True)
    total_cell.border = THIN_BORDER
    total_val_cell = ws_cost.cell(row=total_row, column=2, value=total_cost)
    total_val_cell.font = Font(bold=True)
    total_val_cell.border = THIN_BORDER
    ws_cost.cell(row=total_row, column=3, value=currency).border = THIN_BORDER

    if cost_data_rows:
        ws_cost.auto_filter.ref = f"A1:C{len(cost_data_rows) + 1}"
    _auto_fit_columns(ws_cost)

    wb.save(output_path)
    log.info("Excel workbook saved: %s", output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Discover Azure resources and costs, then export to Excel.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--confirm",
        action="store_true",
        help="Required flag to authorise execution (reads live subscription data).",
    )
    parser.add_argument(
        "--output",
        default=os.environ.get("AZURE_OUTPUT_PATH", "azure_discovery.xlsx"),
        help="Path for the output Excel file (default: azure_discovery.xlsx).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    if not args.confirm:
        print(
            "[ERROR] --confirm flag is required to execute this script.\n"
            "        This script makes live API calls to your Azure subscription.\n"
            "        Re-run with: python agent_d.py --confirm",
            file=sys.stderr,
        )
        return 3

    # ── Credentials ──────────────────────────────────────────────────────
    subscription_id = _env("AZURE_SUBSCRIPTION_ID")

    try:
        credential = build_credential()
    except Exception as exc:  # noqa: BLE001
        log.error("Failed to build Azure credential: %s", exc)
        return 1

    # ── Clients ───────────────────────────────────────────────────────────
    try:
        resource_client = ResourceManagementClient(credential, subscription_id)
        cost_client = CostManagementClient(credential)
    except Exception as exc:  # noqa: BLE001
        log.error("Failed to initialise Azure SDK clients: %s", exc)
        return 1

    # ── Discovery ─────────────────────────────────────────────────────────
    try:
        resource_groups = list_resource_groups(resource_client)
    except Exception as exc:  # noqa: BLE001
        log.error("Resource group discovery failed: %s", exc)
        return 2

    try:
        resources = list_resources(resource_client, subscription_id)
    except Exception as exc:  # noqa: BLE001
        log.error("Resource listing failed: %s", exc)
        return 2

    try:
        total_cost, currency, cost_rows = get_current_month_costs(cost_client, subscription_id)
    except Exception as exc:  # noqa: BLE001
        log.warning("Cost retrieval failed — continuing without cost data: %s", exc)
        total_cost, currency, cost_rows = 0.0, "N/A", []

    # ── Report ────────────────────────────────────────────────────────────
    try:
        generate_excel(
            subscription_id=subscription_id,
            resource_groups=resource_groups,
            resources=resources,
            total_cost=total_cost,
            currency=currency,
            cost_rows=cost_rows,
            output_path=args.output,
        )
    except Exception as exc:  # noqa: BLE001
        log.error("Excel generation failed: %s", exc)
        return 2

    log.info(
        "Discovery complete. Resource groups: %d | Resources: %d | "
        "Cost (%s): %.2f | Output: %s",
        len(resource_groups), len(resources), currency, total_cost, args.output,
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())